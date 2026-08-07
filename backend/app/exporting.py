"""Shared helpers to export survey responses to CSV / XLSX."""

import io
import json
import re
from typing import List

from app.models import Survey, SurveyResponse

_META = [
    "id", "submitted_at", "completed", "respondent_name", "respondent_email",
    "score", "max_score", "percent", "needs_review",
]


def survey_columns(schema: dict) -> List[tuple]:
    """(name, title) for real questions, skipping companion media / display els."""
    cols: List[tuple] = []
    for page in (schema or {}).get("pages", []) or []:
        for el in page.get("elements", []) or []:
            name = el.get("name")
            if not name or name.endswith("__img") or name.endswith("__vid"):
                continue
            if el.get("type") in ("html", "image", "expression"):
                continue
            cols.append((name, el.get("title") or name))
    return cols


def cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def export_rows(survey: Survey, responses: List[SurveyResponse]) -> List[list]:
    from app.identity import identity_fields, respondent_identity

    schema = survey.json_schema or {}
    cols = survey_columns(schema)
    # Quién respondió: el invitado si lo hay; si no, lo que contestó (nombre/mail).
    fields = identity_fields(schema)
    rows: List[list] = [_META + [title for (_n, title) in cols]]
    for r in responses:
        answers = r.answers or {}
        grade = r.grade or {}
        ans_name, ans_email = respondent_identity(schema, answers, fields=fields)
        rows.append(
            [
                str(r.id),
                r.submitted_at.isoformat() if r.submitted_at else "",
                "sí" if r.completed else "no",
                cell(ans_name or ""),
                cell(r.respondent_email or ans_email or ""),
                cell(r.score), cell(r.max_score), cell(grade.get("percent")),
                "sí" if r.needs_review else "no",
            ]
            + [cell(answers.get(name)) for (name, _t) in cols]
        )
    return rows


def rows_to_csv(rows: List[list]) -> bytes:
    import csv

    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel reads UTF-8


def _safe_sheet_name(name: str, used: set) -> str:
    # Excel sheet names: ≤31 chars, no []:*?/\
    clean = re.sub(r"[\[\]:*?/\\]", " ", name or "Hoja").strip()[:31] or "Hoja"
    base, i = clean, 1
    while clean.lower() in used:
        suffix = f" ({i})"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


def sheets_to_xlsx(sheets: List[tuple]) -> bytes:
    """sheets: list of (title, rows). One worksheet per entry."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    used: set = set()
    if not sheets:
        wb.create_sheet("Vacío")
    for title, rows in sheets:
        ws = wb.create_sheet(_safe_sheet_name(title, used))
        for row in rows or [[]]:
            ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MEDIA = "text/csv; charset=utf-8"
